from openpyxl import load_workbook
import pandas as pd

import dash
import dash_core_components as dcc
import dash_html_components as html

import sys


#wb = load_workbook(filename = 'dummyData.xlsx')
#sheet_ranges = wb['Sheet1'] #use sheetname as table name
#print(sheet_ranges['D4'].value)


def loadData():
    HDB_data = pd.read_csv('data/resale-flat-prices-based-on-registration-date-from-jan-2017-onwards.csv')
    HDB_data['Full_Address'] = 'Block ' + HDB_data['block'].str.cat( HDB_data['street_name'], sep=' ')
    HDB_Address_Coord_Distances= pd.read_csv('data/HDB_Address_Coord_Distances.csv')
    complete_hdb_data = HDB_data.set_index('Full_Address').join(HDB_Address_Coord_Distances.set_index('Full_Address'))
    complete_hdb_data.reset_index(inplace = True)

     # Data Cleaning
    complete_hdb_data['resale_price'] = complete_hdb_data['resale_price'].astype('float')
    complete_hdb_data['floor_area_sqm'] = complete_hdb_data['floor_area_sqm'].astype('float')
    complete_hdb_data['lease_commence_date'] = complete_hdb_data['lease_commence_date'].astype('int64')
    complete_hdb_data['lease_remain_years'] = 2020 - complete_hdb_data['lease_commence_date']
    complete_hdb_data['price_per_sqm'] = complete_hdb_data['resale_price'].div(complete_hdb_data['floor_area_sqm'])
    # Let's examine the correlation between our features
    df_numerical_cols = complete_hdb_data[['min_dist_mrt', 'min_dist_mall', 'cbd_dist','floor_area_sqm', 'lease_remain_years', 'resale_price', 'price_per_sqm']]

    df_numerical_cols.describe().round(0)

    df_numerical_cols.dropna(inplace=True)

    # Separate our numerical and categorical variables
    cat_features = ['town', 'flat_type', 'storey_range']
    num_features = ['min_dist_mrt', 'min_dist_mall', 'cbd_dist','floor_area_sqm', 'lease_remain_years', 'resale_price', 'price_per_sqm']
    target = ['resale_price']
    # New data frame for our regression model
    df_reg = complete_hdb_data[['town', 'flat_type', 'storey_range', 'min_dist_mrt', 'min_dist_mall', 'cbd_dist','floor_area_sqm', 'lease_remain_years', 'resale_price', 'price_per_sqm']]
    flat_type_map = {
    'EXECUTIVE': 7,
    'MULTI-GENERATION': 6,
    '5 ROOM': 5,
    '4 ROOM': 4,
    '3 ROOM': 3,
    '2 ROOM': 2,
    '1 ROOM': 1
    }
    df_reg['storey_mean'] = df_reg['storey_range'].apply(lambda x: split_mean(x))
    df_reg = pd.get_dummies(data=df_reg, columns=['town'], drop_first=True)
    print(df_reg)
def split_mean(x):
    split_list = x.split(' TO ')
    mean = (float(split_list[0])+float(split_list[1]))/2
    return mean

def daoTest():
    print("d")

#ws = wb.active
#first_column = ws['B']

# Print the contents
#for x in range(len(first_column)): 
#    print(first_column[x].value) 


#data = pd.read_excel('dummyData.xlsx', index_col=None, header=0)

#print(data['open'])